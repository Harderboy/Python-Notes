# xlsx xls
# 处理xlsx格式的表格
# openpyxl

from openpyxl.reader.excel import load_workbook


def xlslFileReader(path):
    dict = {}
    # 打开文件
    file = load_workbook(filename=path)
    # 所有表格的名称
    sheets = file.get_sheet_names()

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        # 一张表所有的数据
        sheetInfo = []
        # 读取数据
        for linenum in range(1, sheet.max_row + 1):
            # print(linenum)
            lineList = []
            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(row=linenum, column=column).value
                if value is not None:  # 不要用 value != None
                    lineList.append(value)
            # print(lineList)
            sheetInfo.append(lineList)
            # print(sheetInfo)

        # 将一张表的数据存到字典
        dict[sheetName] = sheetInfo
    # print(dict)
    return dict


if __name__ == "__main__":
    filepath = r"C:\Users\liu heng\Desktop\Python学习笔记\文件处理\记分册_1班.xlsx"
    dict = xlslFileReader(filepath)
    print(dict)
    print(dict["记分册"])
    # print(len(dict))

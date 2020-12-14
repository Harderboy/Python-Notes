# xlsx xls
# 处理xlsx格式的表格
# openpyxl

from openpyxl.reader.excel import load_workbook


def xlslFileReader(path):
    # 打开文件
    file = load_workbook(filename=path)
    # 所有表格的名称
    sheets = file.get_sheet_names()
    # print(sheets)

    # 拿出一个表格
    sheet = file.get_sheet_by_name(sheets[0])
    # 最大行数
    # print(sheet.max_row)
    # 最大列数
    # print(sheet.max_column)
    # 表名
    print(sheet.title)

    # 读取数据
    for linenum in range(1, sheet.max_row + 1):
        # print(linenum)
        lineList = []
        for column in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=linenum, column=column).value
            if value is not None:  # 不要用 value != None
                lineList.append(value)

        print(lineList)


if __name__ == "__main__":
    filepath = r"C:\Users\liu heng\Desktop\Python学习笔记\文件处理\记分册_1班.xlsx"
    xlslFileReader(filepath)
